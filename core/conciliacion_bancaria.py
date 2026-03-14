"""
core/conciliacion_bancaria.py — RPA Suite v5.9
================================================
Parser universal de extractos bancarios argentinos.
Soporta PDF y Excel/CSV de todos los bancos principales.

Bancos soportados:
  Galicia · Santander · BBVA · Nación · Provincia · Macro
  ICBC · Patagonia · Ciudad · Supervielle · Brubank · Naranja X
  Mercado Pago + cualquier formato genérico

Estrategia:
  1. Detectar banco por heurísticas del PDF (texto, logo, header)
  2. Aplicar parser específico del banco
  3. Si falla → parser genérico universal
  4. Normalizar a schema común: fecha|descripcion|debito|credito|saldo
"""
import re
import io
from pathlib import Path
from datetime import datetime, date
from typing import Optional

try:
    import pandas as pd
    _PD = True
except ImportError:
    _PD = False

try:
    import pdfplumber
    _PDF = True
except ImportError:
    _PDF = False

# ══════════════════════════════════════════════════════════════
# SCHEMA NORMALIZADO DE MOVIMIENTO
# ══════════════════════════════════════════════════════════════
# Cada movimiento retorna este dict:
# {
#   "fecha":       date,
#   "descripcion": str,
#   "debito":      float,   # siempre positivo
#   "credito":     float,   # siempre positivo
#   "importe":     float,   # negativo=débito, positivo=crédito
#   "saldo":       float,
#   "referencia":  str,     # nro operación si existe
#   "banco":       str,
#   "cuenta":      str,
# }

LogFunc = Optional[callable]


def _parse_fecha(texto: str) -> Optional[date]:
    """Parsea fechas en múltiples formatos argentinos."""
    if not texto:
        return None
    texto = str(texto).strip()
    formatos = [
        "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y",
        "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%y %H:%M:%S",
        "%d %b %Y", "%d %b %y",
    ]
    meses_es = {"ene":"jan","feb":"feb","mar":"mar","abr":"apr","may":"may",
                "jun":"jun","jul":"jul","ago":"aug","sep":"sep","oct":"oct",
                "nov":"nov","dic":"dec"}
    texto_low = texto.lower()
    for es, en in meses_es.items():
        texto_low = texto_low.replace(es, en)
    for fmt in formatos:
        try:
            return datetime.strptime(texto_low, fmt).date()
        except ValueError:
            pass
    # Extraer solo la parte de fecha si hay hora
    m = re.match(r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})', texto)
    if m:
        return _parse_fecha(m.group(1))
    return None


def _parse_monto(texto: str) -> float:
    """Parsea montos argentinos: $1.234,56 → 1234.56"""
    if texto is None:
        return 0.0
    t = str(texto).strip()
    negativo = t.startswith("-") or t.startswith("(") or "DB" in t.upper()
    t = re.sub(r'[^\d,.]', '', t)
    if not t:
        return 0.0
    # Formato AR: 1.234,56
    if ',' in t and '.' in t:
        t = t.replace('.', '').replace(',', '.')
    elif ',' in t:
        t = t.replace(',', '.')
    try:
        v = float(t)
        return -v if negativo else v
    except ValueError:
        return 0.0


def _limpiar_descripcion(texto: str) -> str:
    """Normaliza descripción removiendo espacios extra y caracteres extraños."""
    if not texto:
        return ""
    return re.sub(r'\s+', ' ', str(texto)).strip()


# ══════════════════════════════════════════════════════════════
# DETECCIÓN DE BANCO
# ══════════════════════════════════════════════════════════════

_FIRMAS_BANCOS = {
    "galicia":       ["galicia", "banco galicia", "bco. galicia"],
    "santander":     ["santander", "banco santander", "santander rio", "rio bank"],
    "bbva":          ["bbva", "banco bbva", "francés", "frances", "bbva frances"],
    "nacion":        ["banco de la nacion", "bna", "banco nacion", "nación argentina"],
    "provincia":     ["banco provincia", "bapro", "banco de la provincia"],
    "macro":         ["banco macro", "bco macro"],
    "icbc":          ["icbc", "industrial and commercial"],
    "patagonia":     ["banco patagonia", "patagonia"],
    "ciudad":        ["banco ciudad", "bco. ciudad", "ciudad de buenos aires"],
    "supervielle":   ["supervielle", "banco supervielle"],
    "brubank":       ["brubank"],
    "naranjax":      ["naranja x", "naranjax", "naranja"],
    "mercadopago":   ["mercado pago", "mercadopago", "mp cuenta"],
}

def _detectar_banco(texto_pdf: str) -> str:
    """Detecta el banco a partir del texto extraído del PDF."""
    texto_low = texto_pdf.lower()[:3000]
    for banco, firmas in _FIRMAS_BANCOS.items():
        if any(f in texto_low for f in firmas):
            return banco
    return "generico"


# ══════════════════════════════════════════════════════════════
# PARSERS ESPECÍFICOS POR BANCO
# ══════════════════════════════════════════════════════════════

def _parsear_tabla_vtex(tabla: list, banco: str, cuenta: str) -> list[dict]:
    """
    Parser genérico para tablas estilo VTEX/columnar.
    Detecta automáticamente qué columna es cada campo.
    """
    if not tabla or len(tabla) < 2:
        return []

    # Detectar header
    header = [str(c).lower().strip() if c else "" for c in tabla[0]]

    idx = {
        "fecha":    next((i for i, h in enumerate(header) if any(k in h for k in
                          ["fecha","date","fec"])), None),
        "desc":     next((i for i, h in enumerate(header) if any(k in h for k in
                          ["descripcion","descripción","concepto","detalle","movimiento","referencia"])), None),
        "debito":   next((i for i, h in enumerate(header) if any(k in h for k in
                          ["debito","débito","debe","cargo","egreso","salida"])), None),
        "credito":  next((i for i, h in enumerate(header) if any(k in h for k in
                          ["credito","crédito","haber","abono","ingreso","entrada"])), None),
        "importe":  next((i for i, h in enumerate(header) if any(k in h for k in
                          ["importe","monto","impo"])), None),
        "saldo":    next((i for i, h in enumerate(header) if "saldo" in h), None),
    }

    movimientos = []
    for fila in tabla[1:]:
        if not fila or all(not c for c in fila):
            continue

        fecha_raw = fila[idx["fecha"]] if idx["fecha"] is not None and idx["fecha"] < len(fila) else None
        fecha     = _parse_fecha(str(fecha_raw) if fecha_raw else "")
        if not fecha:
            continue

        desc = _limpiar_descripcion(
            fila[idx["desc"]] if idx["desc"] is not None and idx["desc"] < len(fila) else ""
        )

        # Montos
        if idx["debito"] is not None and idx["credito"] is not None:
            deb = abs(_parse_monto(fila[idx["debito"]] if idx["debito"] < len(fila) else ""))
            cre = abs(_parse_monto(fila[idx["credito"]] if idx["credito"] < len(fila) else ""))
            imp = cre - deb
        elif idx["importe"] is not None:
            imp = _parse_monto(fila[idx["importe"]] if idx["importe"] < len(fila) else "")
            deb = abs(imp) if imp < 0 else 0.0
            cre = imp if imp > 0 else 0.0
        else:
            # Último recurso: buscar columnas numéricas
            numericos = []
            for i, v in enumerate(fila):
                m = _parse_monto(str(v) if v else "")
                if m != 0:
                    numericos.append((i, m))
            if not numericos:
                continue
            imp = numericos[0][1]
            deb = abs(imp) if imp < 0 else 0.0
            cre = imp if imp > 0 else 0.0

        saldo = _parse_monto(fila[idx["saldo"]] if idx["saldo"] is not None and idx["saldo"] < len(fila) else "")

        movimientos.append({
            "fecha":       fecha,
            "descripcion": desc,
            "debito":      round(deb, 2),
            "credito":     round(cre, 2),
            "importe":     round(imp, 2),
            "saldo":       round(saldo, 2),
            "referencia":  "",
            "banco":       banco,
            "cuenta":      cuenta,
        })

    return movimientos


def _parsear_galicia(paginas: list, cuenta: str) -> list[dict]:
    """
    Galicia: tabla con columnas Fecha | Descripción | Débitos | Créditos | Saldo
    A veces usa 2 columnas de importe con signo.
    """
    movimientos = []
    for page in paginas:
        for tabla in (page.extract_tables() or []):
            if not tabla:
                continue
            # Galicia usa header en primera fila
            header_raw = [str(c).lower() if c else "" for c in tabla[0]]
            if not any("fecha" in h or "date" in h for h in header_raw):
                continue
            movimientos.extend(_parsear_tabla_vtex(tabla, "Galicia", cuenta))
    return movimientos


def _parsear_santander(paginas: list, cuenta: str) -> list[dict]:
    """Santander: Fecha | Descripción | Importe | Saldo — importe con signo."""
    movimientos = []
    for page in paginas:
        for tabla in (page.extract_tables() or []):
            movimientos.extend(_parsear_tabla_vtex(tabla, "Santander", cuenta))
    return movimientos


def _parsear_bbva(paginas: list, cuenta: str) -> list[dict]:
    """BBVA (ex Francés): Fecha | Movimiento | Debe | Haber | Saldo"""
    movimientos = []
    for page in paginas:
        for tabla in (page.extract_tables() or []):
            movimientos.extend(_parsear_tabla_vtex(tabla, "BBVA", cuenta))
    return movimientos


def _parsear_nacion(paginas: list, cuenta: str) -> list[dict]:
    """
    Banco Nación: formato más irregular, a veces texto plano.
    Patrón: DD/MM/YYYY  DESCRIPCION  MONTO  SALDO
    """
    movimientos = []
    patron = re.compile(
        r'(\d{2}/\d{2}/\d{4})\s+'           # fecha
        r'(.{5,60?}?)\s+'                    # descripcion (non-greedy)
        r'([\d\.,]+)\s+'                     # monto
        r'([\d\.,]+)'                        # saldo
    )
    for page in paginas:
        # Primero intentar tablas
        for tabla in (page.extract_tables() or []):
            movs = _parsear_tabla_vtex(tabla, "Nación", cuenta)
            if movs:
                movimientos.extend(movs)
                continue
        # Fallback texto
        texto = page.extract_text() or ""
        for m in patron.finditer(texto):
            fecha = _parse_fecha(m.group(1))
            if not fecha:
                continue
            imp   = _parse_monto(m.group(3))
            saldo = _parse_monto(m.group(4))
            movimientos.append({
                "fecha": fecha, "descripcion": _limpiar_descripcion(m.group(2)),
                "debito": abs(imp) if imp < 0 else 0.0,
                "credito": imp if imp > 0 else 0.0,
                "importe": imp, "saldo": saldo,
                "referencia": "", "banco": "Nación", "cuenta": cuenta,
            })
    return movimientos


def _parsear_mercadopago(df: "pd.DataFrame", cuenta: str) -> list[dict]:
    """
    Mercado Pago: CSV con columnas propias.
    Columnas típicas: FECHA, DESCRIPCION, TIPO, MONTO, SALDO
    """
    cols_lower = {c.lower().strip(): c for c in df.columns}
    col_fecha = next((cols_lower[k] for k in ["fecha","date","fecha operacion"] if k in cols_lower), df.columns[0])
    col_desc  = next((cols_lower[k] for k in ["descripcion","descripción","detalle","concepto"] if k in cols_lower), df.columns[1] if len(df.columns)>1 else None)
    col_monto = next((cols_lower[k] for k in ["monto","importe","amount","total"] if k in cols_lower), None)
    col_saldo = next((cols_lower[k] for k in ["saldo","balance"] if k in cols_lower), None)

    movimientos = []
    for _, row in df.iterrows():
        fecha = _parse_fecha(str(row[col_fecha]))
        if not fecha:
            continue
        desc  = _limpiar_descripcion(row[col_desc]) if col_desc else ""
        imp   = _parse_monto(row[col_monto]) if col_monto else 0.0
        saldo = _parse_monto(row[col_saldo]) if col_saldo else 0.0
        movimientos.append({
            "fecha": fecha, "descripcion": desc,
            "debito":  abs(imp) if imp < 0 else 0.0,
            "credito": imp if imp > 0 else 0.0,
            "importe": imp, "saldo": saldo,
            "referencia": "", "banco": "MercadoPago", "cuenta": cuenta,
        })
    return movimientos


# ── Dispatcher de parsers ────────────────────────────────────
_PARSERS_PDF = {
    "galicia":     _parsear_galicia,
    "santander":   _parsear_santander,
    "bbva":        _parsear_bbva,
    "nacion":      _parsear_nacion,
    # Para el resto usamos el genérico
}


# ══════════════════════════════════════════════════════════════
# PARSER GENÉRICO UNIVERSAL (fallback)
# ══════════════════════════════════════════════════════════════

def _parsear_generico_pdf(paginas: list, banco: str, cuenta: str) -> list[dict]:
    """
    Parser universal que:
    1. Extrae todas las tablas de cada página
    2. Identifica la tabla que parece un extracto bancario
    3. Aplica _parsear_tabla_vtex
    """
    movimientos = []
    for page in paginas:
        tablas = page.extract_tables() or []
        for tabla in tablas:
            if not tabla or len(tabla) < 3:
                continue
            # ¿Parece un extracto? Buscar fila con 'fecha'
            header = [str(c).lower() if c else "" for c in tabla[0]]
            if any(k in " ".join(header) for k in ["fecha","date","fec"]):
                movs = _parsear_tabla_vtex(tabla, banco, cuenta)
                movimientos.extend(movs)
                continue
            # Si no hay header obvio, intentar detectar por contenido
            # (primera columna con fechas = es un extracto)
            fechas_col0 = sum(1 for row in tabla[1:6]
                              if row and _parse_fecha(str(row[0] or "")))
            if fechas_col0 >= 2:
                # Inyectar header genérico
                tabla_con_header = [["fecha","descripcion","debito","credito","saldo"]] + tabla
                movs = _parsear_tabla_vtex(tabla_con_header, banco, cuenta)
                movimientos.extend(movs)

        # Fallback texto plano si no hubo tablas
        if not movimientos:
            texto = page.extract_text() or ""
            patron = re.compile(
                r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})'   # fecha
                r'\s+(.{3,60?}?)\s+'                     # descripcion
                r'([\d\.,]+(?:,\d{2})?)'                 # monto
                r'(?:\s+([\d\.,]+(?:,\d{2})?))?'         # saldo opcional
            )
            for m in patron.finditer(texto):
                fecha = _parse_fecha(m.group(1))
                if not fecha:
                    continue
                imp   = _parse_monto(m.group(3))
                saldo = _parse_monto(m.group(4)) if m.group(4) else 0.0
                movimientos.append({
                    "fecha": fecha,
                    "descripcion": _limpiar_descripcion(m.group(2)),
                    "debito":  abs(imp) if imp < 0 else 0.0,
                    "credito": imp if imp > 0 else 0.0,
                    "importe": imp, "saldo": saldo,
                    "referencia": "", "banco": banco, "cuenta": cuenta,
                })

    return movimientos


# ══════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL — LEER EXTRACTO
# ══════════════════════════════════════════════════════════════

def leer_extracto(
    ruta: str,
    banco_forzado: str = None,
    cuenta: str = "",
    log_func: LogFunc = None,
) -> dict:
    """
    Lee un extracto bancario (PDF, Excel o CSV) y retorna
    dict con movimientos normalizados.

    Args:
        ruta:          Ruta al archivo
        banco_forzado: Forzar banco si la detección falla
        cuenta:        Número/alias de cuenta (opcional)
        log_func:      Función de logging

    Retorna:
        {
          "banco":          str,
          "cuenta":         str,
          "movimientos":    list[dict],   ← schema normalizado
          "df":             pd.DataFrame,
          "periodo_desde":  date,
          "periodo_hasta":  date,
          "total_debitos":  float,
          "total_creditos": float,
          "saldo_final":    float,
          "n_movimientos":  int,
          "error":          str (solo si falla)
        }
    """
    def _log(m):
        if log_func: log_func(m)

    if not _PD:
        return {"error": "pandas no instalado"}

    ruta_p = Path(ruta)
    ext    = ruta_p.suffix.lower()
    _log(f"📂 Leyendo extracto: {ruta_p.name}")

    movimientos = []
    banco_det   = banco_forzado or "generico"

    # ── PDF ─────────────────────────────────────────────────
    if ext == ".pdf":
        if not _PDF:
            return {"error": "pdfplumber no instalado. pip install pdfplumber"}
        try:
            with pdfplumber.open(ruta) as pdf:
                # Texto completo para detección de banco
                texto_total = "\n".join(
                    (p.extract_text() or "") for p in pdf.pages[:3]
                )
                if not banco_forzado:
                    banco_det = _detectar_banco(texto_total)
                _log(f"  🏦 Banco detectado: {banco_det.upper()}")

                # Extraer número de cuenta si existe
                m_cuenta = re.search(r'(?:cuenta|cta|N°|nro)[:\s]+([0-9\-/]+)', texto_total, re.I)
                if m_cuenta and not cuenta:
                    cuenta = m_cuenta.group(1).strip()

                paginas = pdf.pages
                parser  = _PARSERS_PDF.get(banco_det, _parsear_generico_pdf)
                movimientos = parser(paginas, banco_det.capitalize(), cuenta)

                # Si el parser específico falló, intentar genérico
                if not movimientos and banco_det != "generico":
                    _log(f"  ⚠ Parser específico sin resultados, usando genérico")
                    movimientos = _parsear_generico_pdf(paginas, banco_det.capitalize(), cuenta)

        except Exception as e:
            return {"error": f"Error leyendo PDF: {e}"}

    # ── Excel / CSV ─────────────────────────────────────────
    elif ext in [".xlsx", ".xls"]:
        try:
            df_raw = pd.read_excel(ruta, engine="openpyxl", header=None)
            # Buscar fila que contenga "fecha"
            header_row = 0
            for i, row in df_raw.iterrows():
                if any("fecha" in str(c).lower() for c in row if c):
                    header_row = i; break
            df_raw = pd.read_excel(ruta, engine="openpyxl", header=header_row)
            banco_det = banco_forzado or _detectar_banco(str(df_raw.to_string()[:2000]))
            if banco_det == "mercadopago":
                movimientos = _parsear_mercadopago(df_raw, cuenta)
            else:
                # Convertir a tabla y usar parser genérico
                tabla = [list(df_raw.columns)] + df_raw.values.tolist()
                movimientos = _parsear_tabla_vtex(tabla, banco_det.capitalize(), cuenta)
        except Exception as e:
            return {"error": f"Error leyendo Excel: {e}"}

    elif ext == ".csv":
        try:
            # Intentar encodings comunes en AR
            for enc in ["utf-8", "latin-1", "iso-8859-1", "cp1252"]:
                try:
                    df_raw = pd.read_csv(ruta, encoding=enc, sep=None, engine="python")
                    break
                except Exception:
                    continue
            banco_det = banco_forzado or _detectar_banco(str(df_raw.columns.tolist()))
            if banco_det == "mercadopago":
                movimientos = _parsear_mercadopago(df_raw, cuenta)
            else:
                tabla = [list(df_raw.columns)] + df_raw.values.tolist()
                movimientos = _parsear_tabla_vtex(tabla, banco_det.capitalize(), cuenta)
        except Exception as e:
            return {"error": f"Error leyendo CSV: {e}"}

    else:
        return {"error": f"Formato no soportado: {ext}. Usá PDF, Excel o CSV."}

    if not movimientos:
        return {"error": "No se encontraron movimientos en el archivo. Verificá que sea un extracto bancario válido."}

    # Ordenar por fecha
    movimientos.sort(key=lambda x: x["fecha"])

    # Construir DataFrame
    df = pd.DataFrame(movimientos)
    df["fecha"] = pd.to_datetime(df["fecha"])

    # Totales
    total_deb = df["debito"].sum()
    total_cre = df["credito"].sum()
    saldo_fin = df["saldo"].iloc[-1] if not df["saldo"].eq(0).all() else (total_cre - total_deb)

    _log(f"  ✅ {len(movimientos)} movimientos | "
         f"Débitos: ${total_deb:,.0f} | Créditos: ${total_cre:,.0f}")

    return {
        "banco":           banco_det.capitalize(),
        "cuenta":          cuenta,
        "movimientos":     movimientos,
        "df":              df,
        "periodo_desde":   movimientos[0]["fecha"],
        "periodo_hasta":   movimientos[-1]["fecha"],
        "total_debitos":   round(total_deb, 2),
        "total_creditos":  round(total_cre, 2),
        "saldo_final":     round(saldo_fin, 2),
        "n_movimientos":   len(movimientos),
    }


# ══════════════════════════════════════════════════════════════
# MOTOR DE CONCILIACIÓN
# ══════════════════════════════════════════════════════════════

def _normalizar_desc(texto: str) -> str:
    """Normaliza descripción para matching: minúsculas, sin acentos, sin puntuación."""
    t = texto.lower()
    for a, b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ñ","n")]:
        t = t.replace(a, b)
    t = re.sub(r'[^\w\s]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def _similitud_movimiento(mov_banco: dict, mov_interno: dict,
                           tolerancia_dias: int = 3,
                           tolerancia_monto_pct: float = 0.01) -> float:
    """
    Calcula score 0-1 de similitud entre movimiento del banco
    y movimiento del registro interno.
    """
    score = 0.0

    # Fecha (hasta 3 días de diferencia)
    diff_dias = abs((mov_banco["fecha"] - mov_interno["fecha"]).days)
    if diff_dias == 0:
        score += 0.5
    elif diff_dias <= tolerancia_dias:
        score += 0.5 * (1 - diff_dias / tolerancia_dias)

    # Monto
    m_banco    = abs(mov_banco.get("importe", 0))
    m_interno  = abs(mov_interno.get("importe", 0))
    if m_banco > 0 and m_interno > 0:
        diff_pct = abs(m_banco - m_interno) / max(m_banco, m_interno)
        if diff_pct <= tolerancia_monto_pct:
            score += 0.4
        elif diff_pct <= 0.05:
            score += 0.2

    # Descripción (palabras en común)
    desc_b = set(_normalizar_desc(mov_banco.get("descripcion","")).split())
    desc_i = set(_normalizar_desc(mov_interno.get("descripcion","")).split())
    if desc_b and desc_i:
        comunes = len(desc_b & desc_i)
        score += 0.1 * min(1, comunes / 2)

    return round(score, 3)


def conciliar(
    extracto: dict,
    movimientos_internos: list[dict],
    umbral_match: float = 0.7,
    tolerancia_dias: int = 3,
    log_func: LogFunc = None,
) -> dict:
    """
    Concilia el extracto bancario contra los movimientos internos.

    movimientos_internos: lista de dicts con keys:
      fecha (date), descripcion (str), importe (float)
      Puede venir del robot de cheques, CxC, o cualquier registro.

    Retorna:
      {
        "conciliados":    list — pares (banco, interno) con score
        "solo_banco":     list — en banco pero no en registros internos
        "solo_interno":   list — en registros internos pero no en banco
        "diferencias":    list — match parcial con diferencia de monto
        "resumen":        dict
      }
    """
    def _log(m):
        if log_func: log_func(m)

    if "error" in extracto:
        return {"error": extracto["error"]}

    movs_banco   = list(extracto.get("movimientos", []))
    movs_interno = list(movimientos_internos)

    _log(f"🔄 Conciliando {len(movs_banco)} movimientos bancarios vs "
         f"{len(movs_interno)} registros internos...")

    conciliados  = []
    solo_banco   = []
    diferencias  = []
    usados_int   = set()

    for i, mb in enumerate(movs_banco):
        mejor_score = 0
        mejor_idx   = None
        for j, mi in enumerate(movs_interno):
            if j in usados_int:
                continue
            score = _similitud_movimiento(mb, mi, tolerancia_dias)
            if score > mejor_score:
                mejor_score = score
                mejor_idx   = j

        if mejor_idx is not None and mejor_score >= umbral_match:
            mi = movs_interno[mejor_idx]
            diff_monto = abs(mb.get("importe",0)) - abs(mi.get("importe",0))
            conciliados.append({
                "banco":     mb,
                "interno":   mi,
                "score":     mejor_score,
                "diff_$":    round(diff_monto, 2),
                "estado":    "✅ Conciliado" if abs(diff_monto) < 1 else "⚠️ Dif. monto",
            })
            usados_int.add(mejor_idx)
            if abs(diff_monto) >= 1:
                diferencias.append(conciliados[-1])
        else:
            solo_banco.append(mb)

    solo_interno = [movs_interno[j] for j in range(len(movs_interno)) if j not in usados_int]

    pct_conc = len(conciliados) / len(movs_banco) * 100 if movs_banco else 0
    _log(f"  ✅ Conciliados: {len(conciliados)} ({pct_conc:.0f}%)")
    _log(f"  🔵 Solo en banco: {len(solo_banco)}")
    _log(f"  🟡 Solo en registros internos: {len(solo_interno)}")
    _log(f"  ⚠️  Con diferencia de monto: {len(diferencias)}")

    return {
        "conciliados":   conciliados,
        "solo_banco":    solo_banco,
        "solo_interno":  solo_interno,
        "diferencias":   diferencias,
        "resumen": {
            "total_banco":      len(movs_banco),
            "total_interno":    len(movs_interno),
            "conciliados":      len(conciliados),
            "solo_banco":       len(solo_banco),
            "solo_interno":     len(solo_interno),
            "con_diferencia":   len(diferencias),
            "pct_conciliado":   round(pct_conc, 1),
        },
    }


# ══════════════════════════════════════════════════════════════
# EXPORTAR INFORME
# ══════════════════════════════════════════════════════════════

def exportar_conciliacion_excel(
    resultado_conciliacion: dict,
    extracto: dict,
    log_func: LogFunc = None,
) -> bytes:
    """Genera Excel con 4 hojas: Resumen, Conciliados, Solo Banco, Solo Interno."""
    def _log(m):
        if log_func: log_func(m)

    if not _PD or "error" in resultado_conciliacion:
        return b""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb  = Workbook()
        res = resultado_conciliacion["resumen"]

        AZUL   = PatternFill("solid", fgColor="1D3557")
        VERDE  = PatternFill("solid", fgColor="D1FAE5")
        ROJO   = PatternFill("solid", fgColor="FEE2E2")
        AMARILLO=PatternFill("solid", fgColor="FEF3C7")
        W_FONT = Font(color="FFFFFF", bold=True)

        def _auto_width(ws):
            for col in ws.columns:
                w = max(len(str(c.value or "")) for c in col) + 3
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(w, 50)

        def _header_row(ws, cols):
            for j, c in enumerate(cols, 1):
                cell = ws.cell(row=ws.max_row, column=j, value=c)
                cell.fill = AZUL; cell.font = W_FONT
                cell.alignment = Alignment(horizontal="center")

        # ── Hoja 1: Resumen ──
        ws1 = wb.active; ws1.title = "Resumen"
        ws1["A1"] = f"Conciliación Bancaria — {extracto.get('banco','')} Cta. {extracto.get('cuenta','')}"
        ws1["A1"].font = Font(bold=True, size=14, color="1D3557")
        ws1.append([])
        for k, v in res.items():
            ws1.append([k.replace("_", " ").title(), v])
        ws1.append([])
        ws1.append(["Período", f"{extracto.get('periodo_desde','')} → {extracto.get('periodo_hasta','')}"])
        ws1.append(["Total débitos",  extracto.get("total_debitos",0)])
        ws1.append(["Total créditos", extracto.get("total_creditos",0)])
        ws1.append(["Saldo final",    extracto.get("saldo_final",0)])
        _auto_width(ws1)

        # ── Hoja 2: Conciliados ──
        ws2 = wb.create_sheet("✅ Conciliados")
        ws2.append(["Fecha Banco","Descripción Banco","Importe Banco",
                    "Fecha Interno","Descripción Interno","Importe Interno",
                    "Diferencia $","Score","Estado"])
        _header_row(ws2, ws2[1])
        for par in resultado_conciliacion["conciliados"]:
            mb = par["banco"]; mi = par["interno"]
            row = ws2.append([
                str(mb["fecha"]), mb["descripcion"], mb["importe"],
                str(mi.get("fecha","")), mi.get("descripcion",""), mi.get("importe",0),
                par["diff_$"], par["score"], par["estado"]
            ])
            fill = VERDE if abs(par["diff_$"]) < 1 else AMARILLO
            for cell in ws2[ws2.max_row]:
                cell.fill = fill
        _auto_width(ws2)

        # ── Hoja 3: Solo en banco ──
        ws3 = wb.create_sheet("🔵 Solo Banco")
        ws3.append(["Fecha","Descripción","Débito","Crédito","Importe","Saldo"])
        _header_row(ws3, ws3[1])
        for mb in resultado_conciliacion["solo_banco"]:
            ws3.append([str(mb["fecha"]), mb["descripcion"],
                        mb["debito"], mb["credito"], mb["importe"], mb["saldo"]])
            for cell in ws3[ws3.max_row]: cell.fill = ROJO
        _auto_width(ws3)

        # ── Hoja 4: Solo en registros internos ──
        ws4 = wb.create_sheet("🟡 Solo Interno")
        ws4.append(["Fecha","Descripción","Importe"])
        _header_row(ws4, ws4[1])
        for mi in resultado_conciliacion["solo_interno"]:
            ws4.append([str(mi.get("fecha","")), mi.get("descripcion",""), mi.get("importe",0)])
            for cell in ws4[ws4.max_row]: cell.fill = AMARILLO
        _auto_width(ws4)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        _log("  ✅ Excel de conciliación generado")
        return buf.getvalue()

    except Exception as e:
        _log(f"  ❌ Error generando Excel: {e}")
        return b""
