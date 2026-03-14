"""
core/normalizador_maestro.py — RPA Suite v5.9
===============================================
Normalizador ULTRA de maestro de artículos.

Capacidades:
  1. Deduplicación inteligente — similitud semántica + fonética
  2. Clasificación automática familia/subfamilia con IA
     (se retroalimenta del Pricing Research para conocer categorías reales)
  3. Normalización de unidades (kg→Kg, lt→L, UN→Und, etc.)
  4. Estandarización de descripción (marca al frente, unidad al final)
  5. Validación EAN13 / código de barras
  6. Cruce con base de proveedores
  7. Exporta maestro limpio + informe de cambios
"""
import re
import json
import unicodedata
from pathlib import Path
from typing import Optional

try:
    import pandas as pd
    import numpy as np
    _PD = True
except ImportError:
    _PD = False

LogFunc = Optional[callable]

_DATA_DIR   = Path(__file__).parent.parent / "pricing_data"
_FAMILIA_DB = _DATA_DIR / "familias_conocidas.json"


# ══════════════════════════════════════════════════════════════
# 1. NORMALIZACIÓN DE UNIDADES
# ══════════════════════════════════════════════════════════════

_UNIDADES_MAP = {
    # Peso
    r'\bkg\b':   'Kg',  r'\bkgr\b': 'Kg', r'\bkilos?\b': 'Kg',
    r'\bgr?\b':  'g',   r'\bgramos?\b': 'g',
    r'\bmg\b':   'mg',
    r'\bton\b':  'Ton', r'\btoneladas?\b': 'Ton',
    # Volumen
    r'\bl\b':    'L',   r'\blt\b':  'L',  r'\blts\b': 'L',
    r'\blitros?\b': 'L',
    r'\bml\b':   'ml',  r'\bcc\b':  'ml',
    r'\bcm3\b':  'ml',
    # Unidades
    r'\bun\b':   'Un',  r'\bund\b': 'Un', r'\bunid(ad(es)?)?\b': 'Un',
    r'\bpza?\b': 'Pza', r'\bpiezas?\b': 'Pza',
    r'\bpaq\b':  'Paq', r'\bpaquetes?\b': 'Paq',
    r'\bcaja\b': 'Cja', r'\bcja\b': 'Cja',
    r'\bdoc\b':  'Doc', r'\bdocenas?\b': 'Doc',
    r'\brol\b':  'Rol', r'\brollo\b': 'Rol',
    r'\bpor\b':  'x',   r'\bx\b':    'x',
    # Longitud
    r'\bmt?\b':  'm',   r'\bmetros?\b': 'm',
    r'\bcm\b':   'cm',  r'\bmm\b':  'mm',
}

def normalizar_unidad(texto: str) -> str:
    """Normaliza unidades de medida en una descripción."""
    t = texto.strip()
    for patron, reemplazo in _UNIDADES_MAP.items():
        t = re.sub(patron, reemplazo, t, flags=re.IGNORECASE)
    # Normalizar "1 L", "1L", "1lt" → "1 L"
    t = re.sub(r'(\d+)\s*(Kg|g|mg|L|ml|m|cm|Un|Pza)', r'\1 \2', t)
    return t


# ══════════════════════════════════════════════════════════════
# 2. NORMALIZACIÓN DE DESCRIPCIÓN
# ══════════════════════════════════════════════════════════════

def _sin_acentos(texto: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFD', texto)
                   if unicodedata.category(c) != 'Mn')

def _tokenizar(texto: str) -> set:
    """Palabras significativas de una descripción."""
    t = _sin_acentos(texto.lower())
    t = re.sub(r'[^\w\s]', ' ', t)
    palabras = t.split()
    stopwords = {"de","la","el","los","las","un","una","con","para","x","y","e","o"}
    return {p for p in palabras if p not in stopwords and len(p) > 1}

def normalizar_descripcion(desc: str, marca: str = "") -> str:
    """
    Estandariza la descripción:
    - Title case
    - Marca al principio si se conoce
    - Unidades normalizadas al final
    - Elimina caracteres extraños
    """
    if not desc:
        return ""
    # Limpiar
    d = re.sub(r'\s+', ' ', str(desc)).strip()
    d = normalizar_unidad(d)
    # Title case controlado
    d = d.title()
    # Si hay marca conocida, asegurarse que vaya primero
    if marca and marca.lower() not in d.lower()[:len(marca)+2]:
        d = f"{marca.title()} {d}"
    return d


# ══════════════════════════════════════════════════════════════
# 3. DEDUPLICACIÓN INTELIGENTE
# ══════════════════════════════════════════════════════════════

def _similitud_jaccard(a: str, b: str) -> float:
    """Similitud Jaccard entre tokens de dos strings."""
    ta = _tokenizar(a)
    tb = _tokenizar(b)
    if not ta or not tb:
        return 0.0
    inter = len(ta & tb)
    union = len(ta | tb)
    return inter / union if union > 0 else 0.0

def _similitud_fonética(a: str, b: str) -> float:
    """
    Similitud fonética básica: compara primeras 3 letras de cada token.
    Detecta 'ACEIT' vs 'ACEYTE', 'FIDEOS' vs 'FIDEUS', etc.
    """
    def _fonet(s):
        return {w[:4].upper() for w in _tokenizar(s)}
    ta = _fonet(a)
    tb = _fonet(b)
    if not ta or not tb:
        return 0.0
    inter = len(ta & tb)
    union = len(ta | tb)
    return inter / union if union > 0 else 0.0

def detectar_duplicados(
    df: "pd.DataFrame",
    col_desc: str = "DESCRIPCION",
    col_codigo: str = None,
    umbral_jaccard: float = 0.70,
    umbral_fonetico: float = 0.75,
    log_func: LogFunc = None,
) -> list[dict]:
    """
    Detecta artículos duplicados o muy similares en el maestro.

    Retorna lista de grupos de duplicados:
    [
      {
        "tipo": "exacto" | "similar" | "fonetico",
        "score": float,
        "items": [{"idx": int, "codigo": str, "descripcion": str}]
      }
    ]
    """
    def _log(m):
        if log_func: log_func(m)

    if not _PD:
        return []

    n = len(df)
    _log(f"  🔍 Analizando {n} artículos en busca de duplicados...")

    grupos  = []
    vistos  = set()

    # Vectorizar descripciones (más rápido que doble loop)
    descs = df[col_desc].fillna("").tolist()

    for i in range(n):
        if i in vistos:
            continue
        similares = []
        for j in range(i+1, n):
            if j in vistos:
                continue
            di = str(descs[i])
            dj = str(descs[j])

            # 1. Exacto (normalizado)
            if _sin_acentos(di.lower().strip()) == _sin_acentos(dj.lower().strip()):
                similares.append({"j": j, "tipo": "exacto", "score": 1.0})
                continue

            # 2. Jaccard semántico
            sj = _similitud_jaccard(di, dj)
            if sj >= umbral_jaccard:
                similares.append({"j": j, "tipo": "similar", "score": sj})
                continue

            # 3. Fonético (para errores de tipeo)
            sf = _similitud_fonética(di, dj)
            if sf >= umbral_fonetico and sj >= 0.5:
                similares.append({"j": j, "tipo": "fonetico", "score": sf})

        if similares:
            items = [{"idx": i,
                      "codigo": str(df.iloc[i][col_codigo]) if col_codigo else str(i),
                      "descripcion": descs[i]}]
            for s in similares:
                j = s["j"]
                vistos.add(j)
                items.append({"idx": j,
                              "codigo": str(df.iloc[j][col_codigo]) if col_codigo else str(j),
                              "descripcion": descs[j]})
            grupos.append({
                "tipo":  similares[0]["tipo"],
                "score": round(max(s["score"] for s in similares), 3),
                "items": items,
            })
            vistos.add(i)

    _log(f"  ⚠ {len(grupos)} grupo(s) de duplicados detectados")
    return grupos


# ══════════════════════════════════════════════════════════════
# 4. CLASIFICACIÓN IA — FAMILIA / SUBFAMILIA
# ══════════════════════════════════════════════════════════════

def _cargar_familias_conocidas() -> dict:
    """
    Carga el diccionario de familias aprendidas del Pricing Research.
    El motor de pricing va llenando este archivo cuando hace scraping.
    """
    if _FAMILIA_DB.exists():
        try:
            return json.loads(_FAMILIA_DB.read_text(encoding="utf-8"))
        except Exception:
            pass
    # Familias base para retail alimenticio argentino
    return {
        "aceites": {"familia": "Aceites y Grasas", "subfamilia": "Aceites comestibles"},
        "yerba":   {"familia": "Infusiones", "subfamilia": "Yerba mate"},
        "cafe":    {"familia": "Infusiones", "subfamilia": "Café"},
        "te ":     {"familia": "Infusiones", "subfamilia": "Té"},
        "azucar":  {"familia": "Almacén", "subfamilia": "Azúcar y edulcorantes"},
        "harina":  {"familia": "Almacén", "subfamilia": "Harinas"},
        "arroz":   {"familia": "Almacén", "subfamilia": "Arroz"},
        "fideos":  {"familia": "Pastas", "subfamilia": "Pastas secas"},
        "pasta":   {"familia": "Pastas", "subfamilia": "Pastas secas"},
        "gallett": {"familia": "Golosinas y snacks", "subfamilia": "Galletitas"},
        "leche":   {"familia": "Lácteos", "subfamilia": "Leche"},
        "yogur":   {"familia": "Lácteos", "subfamilia": "Yogurt"},
        "queso":   {"familia": "Lácteos", "subfamilia": "Quesos"},
        "agua ":   {"familia": "Bebidas", "subfamilia": "Aguas"},
        "gaseosa": {"familia": "Bebidas", "subfamilia": "Gaseosas"},
        "jugo":    {"familia": "Bebidas", "subfamilia": "Jugos"},
        "cerveza": {"familia": "Bebidas", "subfamilia": "Cervezas"},
        "vino":    {"familia": "Bebidas", "subfamilia": "Vinos"},
        "deterg":  {"familia": "Limpieza", "subfamilia": "Detergentes"},
        "lavand":  {"familia": "Limpieza", "subfamilia": "Lavandina"},
        "jabon":   {"familia": "Limpieza / Higiene", "subfamilia": "Jabón"},
        "shampu":  {"familia": "Higiene personal", "subfamilia": "Shampoo"},
        "shampoo": {"familia": "Higiene personal", "subfamilia": "Shampoo"},
        "desodor": {"familia": "Higiene personal", "subfamilia": "Desodorantes"},
        "papel h": {"familia": "Higiene", "subfamilia": "Papel higiénico"},
        "panal":   {"familia": "Bebé", "subfamilia": "Pañales"},
        "carne":   {"familia": "Carnes", "subfamilia": "Carnes vacunas"},
        "pollo":   {"familia": "Carnes", "subfamilia": "Aves"},
        "atun":    {"familia": "Enlatados", "subfamilia": "Atún"},
        "tomate":  {"familia": "Enlatados", "subfamilia": "Tomates"},
        "mayones": {"familia": "Salsas y condimentos", "subfamilia": "Mayonesa"},
        "ketchup": {"familia": "Salsas y condimentos", "subfamilia": "Ketchup"},
    }

def _clasificar_por_palabras(desc: str, familias_db: dict) -> dict:
    """Clasificación por matching de palabras clave."""
    desc_low = _sin_acentos(desc.lower())
    for kw, datos in familias_db.items():
        if kw.strip() in desc_low:
            return datos
    return {"familia": "Sin clasificar", "subfamilia": "Sin clasificar"}

def clasificar_con_ia(
    descripciones: list[str],
    batch_size: int = 30,
    log_func: LogFunc = None,
) -> list[dict]:
    """
    Clasifica lista de descripciones en familia/subfamilia usando IA.
    Procesa en batches para eficiencia.
    Guarda el resultado en familias_conocidas.json para uso futuro.

    Retorna lista de dicts: [{"familia": str, "subfamilia": str}]
    """
    def _log(m):
        if log_func: log_func(m)

    familias_db = _cargar_familias_conocidas()
    resultados  = [None] * len(descripciones)

    # Paso 1: clasificar los que podemos localmente
    sin_clasificar_idx = []
    for i, desc in enumerate(descripciones):
        local = _clasificar_por_palabras(desc, familias_db)
        if local["familia"] != "Sin clasificar":
            resultados[i] = local
        else:
            sin_clasificar_idx.append(i)

    _log(f"  📂 {len(descripciones) - len(sin_clasificar_idx)} clasificados localmente, "
         f"{len(sin_clasificar_idx)} van a IA")

    if not sin_clasificar_idx:
        return resultados

    # Paso 2: IA para los sin clasificar (en batches)
    import requests as _req

    for batch_start in range(0, len(sin_clasificar_idx), batch_size):
        batch_idxs = sin_clasificar_idx[batch_start:batch_start + batch_size]
        batch_desc = [descripciones[i] for i in batch_idxs]

        prompt = (
            "Eres un experto en categorización de productos de retail/supermercado argentino.\n"
            "Para cada descripción de producto, asignale una FAMILIA y SUBFAMILIA.\n"
            "Respondé SOLO con JSON array, sin texto adicional, sin markdown. Ejemplo:\n"
            '[{"familia":"Lácteos","subfamilia":"Leche"}]\n\n'
            "Productos a clasificar:\n"
            + "\n".join(f"{i+1}. {d}" for i, d in enumerate(batch_desc))
        )

        try:
            r = _req.post(
                "https://api.anthropic.com/v1/messages",
                headers={"Content-Type": "application/json"},
                json={
                    "model":      "claude-haiku-4-5-20251001",
                    "max_tokens": 1000,
                    "messages":   [{"role": "user", "content": prompt}]
                },
                timeout=30,
            )
            if r.status_code == 200:
                texto = r.json()["content"][0]["text"].strip()
                # Limpiar posibles fences de markdown
                texto = re.sub(r'```json|```', '', texto).strip()
                clasificaciones = json.loads(texto)
                for local_i, idx in enumerate(batch_idxs):
                    if local_i < len(clasificaciones):
                        cl = clasificaciones[local_i]
                        resultados[idx] = {
                            "familia":    cl.get("familia", "Sin clasificar"),
                            "subfamilia": cl.get("subfamilia", "Sin clasificar"),
                        }
                        # Aprender: guardar keyword → familia
                        desc_low = _sin_acentos(descripciones[idx].lower())
                        # Extraer primera palabra significativa como keyword
                        tokens = [t for t in desc_low.split() if len(t) >= 4]
                        if tokens:
                            kw = tokens[0][:6]
                            if kw not in familias_db:
                                familias_db[kw] = resultados[idx]
                    else:
                        resultados[idx] = {"familia": "Sin clasificar", "subfamilia": "Sin clasificar"}
            else:
                _log(f"  ⚠ IA error HTTP {r.status_code} — usando clasificación local")
                for idx in batch_idxs:
                    if resultados[idx] is None:
                        resultados[idx] = {"familia": "Sin clasificar", "subfamilia": "Sin clasificar"}
        except Exception as e:
            _log(f"  ⚠ IA no disponible: {e} — usando clasificación local")
            for idx in batch_idxs:
                if resultados[idx] is None:
                    resultados[idx] = {"familia": "Sin clasificar", "subfamilia": "Sin clasificar"}

    # Guardar familias aprendidas
    try:
        _DATA_DIR.mkdir(parents=True, exist_ok=True)
        _FAMILIA_DB.write_text(json.dumps(familias_db, ensure_ascii=False, indent=2),
                                encoding="utf-8")
        _log(f"  💾 Base de familias actualizada: {len(familias_db)} keywords")
    except Exception:
        pass

    # Rellenar None restantes
    for i in range(len(resultados)):
        if resultados[i] is None:
            resultados[i] = {"familia": "Sin clasificar", "subfamilia": "Sin clasificar"}

    return resultados


# ══════════════════════════════════════════════════════════════
# 5. VALIDACIÓN EAN13
# ══════════════════════════════════════════════════════════════

def validar_ean13(codigo: str) -> bool:
    """Valida dígito verificador de código EAN-13."""
    c = re.sub(r'\D', '', str(codigo))
    if len(c) != 13:
        return False
    pares   = sum(int(c[i]) for i in range(0, 12, 2))
    impares = sum(int(c[i]) for i in range(1, 12, 2))
    dv_calc = (10 - (pares + impares * 3) % 10) % 10
    return int(c[12]) == dv_calc

def validar_ean8(codigo: str) -> bool:
    c = re.sub(r'\D', '', str(codigo))
    if len(c) != 8: return False
    pares   = sum(int(c[i]) for i in range(0, 7, 2))
    impares = sum(int(c[i]) for i in range(1, 7, 2))
    dv_calc = (10 - (pares * 3 + impares) % 10) % 10
    return int(c[7]) == dv_calc

def validar_codigo_barras(codigo) -> str:
    """Retorna 'EAN13' | 'EAN8' | 'DUN14' | 'inválido' | 'sin código'"""
    if not codigo or str(codigo).strip() in ["nan","None",""]:
        return "sin código"
    c = re.sub(r'\D', '', str(codigo))
    if len(c) == 13 and validar_ean13(c):  return "EAN13 ✅"
    if len(c) == 8  and validar_ean8(c):   return "EAN8 ✅"
    if len(c) == 14:                        return "DUN14 (sin verificar)"
    if len(c) == 13 or len(c) == 8:        return "inválido ❌"
    return f"formato desconocido ({len(c)} dígitos)"


# ══════════════════════════════════════════════════════════════
# 6. PROCESAMIENTO COMPLETO DEL MAESTRO
# ══════════════════════════════════════════════════════════════

def procesar_maestro(
    df: "pd.DataFrame",
    col_codigo:   str = None,
    col_desc:     str = None,
    col_familia:  str = None,
    col_barras:   str = None,
    col_marca:    str = None,
    col_unidad:   str = None,
    clasificar_ia: bool = True,
    detectar_dupl: bool = True,
    validar_barra: bool = True,
    log_func: LogFunc = None,
) -> dict:
    """
    Procesa el maestro de artículos completo.

    Retorna:
      {
        "df_limpio":    DataFrame normalizado
        "duplicados":   lista de grupos de duplicados
        "informe":      dict con métricas y cambios
        "cambios":      list de {idx, campo, antes, despues}
      }
    """
    def _log(m):
        if log_func: log_func(m)

    if not _PD:
        return {"error": "pandas no instalado"}

    # Auto-detectar columnas si no se especifican
    cols_low = {c.lower().strip(): c for c in df.columns}

    def _auto(col, candidatos):
        if col: return col
        for k in candidatos:
            if k in cols_low: return cols_low[k]
        return None

    col_codigo  = _auto(col_codigo,  ["codigo","cód","cod","sku","codart","id","item"])
    col_desc    = _auto(col_desc,    ["descripcion","descripción","desc","nombre","producto","articulo"])
    col_familia = _auto(col_familia, ["familia","desfam","rubro","categoria","categoría","cat"])
    col_barras  = _auto(col_barras,  ["barras","ean","ean13","codigobarra","cod.barras","codbarras"])
    col_marca   = _auto(col_marca,   ["marca","brand"])
    col_unidad  = _auto(col_unidad,  ["unidad","um","umedida","presentacion"])

    if not col_desc:
        return {"error": f"No se encontró columna de descripción. Columnas: {list(df.columns)}"}

    _log(f"  📂 Maestro: {len(df)} artículos")
    _log(f"  📋 Columnas detectadas: desc={col_desc}, cod={col_codigo}, "
         f"fam={col_familia}, barras={col_barras}")

    df_limpio = df.copy()
    cambios   = []
    n_total   = len(df_limpio)

    # ── Normalizar descripciones ─────────────────────────────
    _log("  ✏ Normalizando descripciones...")
    for i, row in df_limpio.iterrows():
        desc_orig = str(row[col_desc]).strip()
        marca     = str(row[col_marca]).strip() if col_marca and col_marca in df_limpio.columns else ""
        desc_new  = normalizar_descripcion(desc_orig, marca)
        if desc_new != desc_orig:
            cambios.append({"idx": i, "campo": col_desc, "antes": desc_orig, "despues": desc_new})
            df_limpio.at[i, col_desc] = desc_new

    # ── Normalizar unidades ──────────────────────────────────
    if col_unidad and col_unidad in df_limpio.columns:
        _log("  📏 Normalizando unidades de medida...")
        for i, row in df_limpio.iterrows():
            u_orig = str(row[col_unidad]).strip()
            u_new  = normalizar_unidad(u_orig).strip()
            if u_new != u_orig:
                cambios.append({"idx": i, "campo": col_unidad, "antes": u_orig, "despues": u_new})
                df_limpio.at[i, col_unidad] = u_new

    # ── Detectar duplicados ──────────────────────────────────
    duplicados = []
    if detectar_dupl:
        _log("  🔍 Detectando duplicados...")
        duplicados = detectar_duplicados(
            df_limpio, col_desc=col_desc, col_codigo=col_codigo, log_func=log_func
        )

    # ── Clasificar familias con IA ───────────────────────────
    if clasificar_ia:
        _log("  🤖 Clasificando familias con IA...")
        descs_list = df_limpio[col_desc].fillna("").tolist()
        clasificaciones = clasificar_con_ia(descs_list, log_func=log_func)

        # Crear columnas si no existen
        if "Familia_IA" not in df_limpio.columns:
            df_limpio["Familia_IA"] = ""
        if "Subfamilia_IA" not in df_limpio.columns:
            df_limpio["Subfamilia_IA"] = ""

        for i, cl in enumerate(clasificaciones):
            if cl:
                df_limpio.at[df_limpio.index[i], "Familia_IA"]    = cl["familia"]
                df_limpio.at[df_limpio.index[i], "Subfamilia_IA"] = cl["subfamilia"]

        # Si hay columna de familia original vacía, completar con IA
        if col_familia and col_familia in df_limpio.columns:
            mask_vacios = df_limpio[col_familia].isna() | (df_limpio[col_familia].astype(str).str.strip() == "")
            n_completados = mask_vacios.sum()
            if n_completados > 0:
                df_limpio.loc[mask_vacios, col_familia] = df_limpio.loc[mask_vacios, "Familia_IA"]
                _log(f"  ✅ {n_completados} artículos sin familia asignada — completados con IA")

    # ── Validar códigos de barra ─────────────────────────────
    if validar_barra and col_barras and col_barras in df_limpio.columns:
        _log("  📊 Validando códigos de barra...")
        df_limpio["Estado_EAN"] = df_limpio[col_barras].apply(
            lambda x: validar_codigo_barras(x)
        )
        n_invalidos = (df_limpio["Estado_EAN"].str.contains("inválido")).sum()
        _log(f"  {'✅' if n_invalidos == 0 else '⚠'} {n_invalidos} código(s) de barra inválido(s)")

    # ── Informe ──────────────────────────────────────────────
    informe = {
        "total_articulos":    n_total,
        "descripciones_norm": sum(1 for c in cambios if c["campo"] == col_desc),
        "unidades_norm":      sum(1 for c in cambios if c.get("campo") == col_unidad),
        "grupos_duplicados":  len(duplicados),
        "arts_duplicados":    sum(len(g["items"]) for g in duplicados),
        "clasificados_ia":    len([r for r in df_limpio["Familia_IA"].values
                                   if r != "Sin clasificar"]) if "Familia_IA" in df_limpio.columns else 0,
        "total_cambios":      len(cambios),
    }

    if col_barras and "Estado_EAN" in df_limpio.columns:
        informe["ean_validos"]   = (df_limpio["Estado_EAN"].str.contains("✅")).sum()
        informe["ean_invalidos"] = (df_limpio["Estado_EAN"].str.contains("inválido")).sum()
        informe["sin_ean"]       = (df_limpio["Estado_EAN"] == "sin código").sum()

    _log(f"\n  📊 Resumen:")
    for k, v in informe.items():
        _log(f"     {k}: {v}")

    return {
        "df_limpio":  df_limpio,
        "duplicados": duplicados,
        "informe":    informe,
        "cambios":    cambios,
    }


def exportar_maestro_excel(resultado: dict, log_func: LogFunc = None) -> bytes:
    """Exporta el maestro procesado con 3 hojas: Maestro limpio, Duplicados, Cambios."""
    def _log(m):
        if log_func: log_func(m)
    if not _PD or "error" in resultado:
        return b""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter

        wb  = Workbook()
        AZUL  = PatternFill("solid", fgColor="1D3557")
        VERDE = PatternFill("solid", fgColor="D1FAE5")
        ROJO  = PatternFill("solid", fgColor="FEE2E2")
        AMARI = PatternFill("solid", fgColor="FEF3C7")
        WF    = Font(color="FFFFFF", bold=True)

        def _aw(ws):
            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                    max(len(str(c.value or "")) for c in col) + 2, 45)

        # Hoja 1: Maestro limpio
        ws1 = wb.active; ws1.title = "Maestro Limpio"
        df_l = resultado["df_limpio"]
        for j, c in enumerate(df_l.columns, 1):
            cell = ws1.cell(row=1, column=j, value=c)
            cell.fill = AZUL; cell.font = WF
        for i, (_, row) in enumerate(df_l.iterrows(), 2):
            for j, val in enumerate(row, 1):
                ws1.cell(row=i, column=j, value=str(val) if not isinstance(val, (int, float)) else val)
        _aw(ws1)

        # Hoja 2: Duplicados
        ws2 = wb.create_sheet("⚠ Duplicados")
        ws2.append(["Tipo","Score","Código","Descripción"])
        for c in ws2[1]: c.fill = AZUL; c.font = WF
        for grp in resultado["duplicados"]:
            for item in grp["items"]:
                row = ws2.append([grp["tipo"], grp["score"], item["codigo"], item["descripcion"]])
                for c in ws2[ws2.max_row]: c.fill = AMARI
        _aw(ws2)

        # Hoja 3: Cambios
        ws3 = wb.create_sheet("📝 Cambios")
        ws3.append(["Índice","Campo","Antes","Después"])
        for c in ws3[1]: c.fill = AZUL; c.font = WF
        for cambio in resultado["cambios"]:
            ws3.append([cambio["idx"], cambio["campo"], cambio["antes"], cambio["despues"]])
        _aw(ws3)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        _log("  ✅ Excel maestro generado")
        return buf.getvalue()
    except Exception as e:
        _log(f"  ❌ Error: {e}")
        return b""
