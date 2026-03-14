"""
dashboards/exportar.py — RPA Suite v5.5
=========================================
Módulo de exportación a Excel y PDF para los dashboards BI.

Funciones:
  generar_excel(df, kpis, titulo, periodo) → bytes
  generar_pdf(df, kpis, titulo, periodo, narrativa_ia) → bytes
"""
import io
from datetime import datetime
from pathlib import Path


# ══════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════

def generar_excel(df=None, kpis: dict = None, titulo: str = "Dashboard BI",
                  periodo: str = "", narrativa_ia: str = "") -> bytes:
    """
    Genera un Excel con múltiples hojas:
      - Resumen: KPIs principales
      - Datos:   DataFrame completo
      - IA:      Narrativa generada por IA (si existe)
    Retorna bytes para st.download_button.
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError(f"openpyxl requerido: pip install openpyxl — {e}")

    wb = Workbook()

    # ── Hoja 1: Resumen KPIs ──────────────────────────────────
    ws_kpi = wb.active
    ws_kpi.title = "Resumen"

    header_fill = PatternFill("solid", fgColor="1D3557")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font  = Font(bold=True, size=14, color="1A1D23")
    alt_fill    = PatternFill("solid", fgColor="EFF6FF")

    ws_kpi["A1"] = titulo
    ws_kpi["A1"].font = title_font
    ws_kpi["A2"] = f"Período: {periodo}" if periodo else f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_kpi["A2"].font = Font(size=10, color="64748B")
    ws_kpi.merge_cells("A1:C1")
    ws_kpi.merge_cells("A2:C2")

    ws_kpi.append([])
    ws_kpi.append(["Indicador", "Valor", "Descripción"])
    hdr_row = ws_kpi.max_row
    for col in range(1, 4):
        cell = ws_kpi.cell(row=hdr_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    if kpis:
        for i, (k, v) in enumerate(kpis.items()):
            desc = ""
            if isinstance(v, dict):
                desc = v.get("desc", "")
                v    = v.get("valor", v)
            row_idx = ws_kpi.max_row + 1
            ws_kpi.append([k, str(v), desc])
            if i % 2 == 0:
                for col in range(1, 4):
                    ws_kpi.cell(row=row_idx, column=col).fill = alt_fill

    ws_kpi.column_dimensions["A"].width = 35
    ws_kpi.column_dimensions["B"].width = 20
    ws_kpi.column_dimensions["C"].width = 40

    # ── Hoja 2: Datos ─────────────────────────────────────────
    if df is not None and not df.empty:
        ws_data = wb.create_sheet("Datos")
        import pandas as pd
        # Headers
        for j, col_name in enumerate(df.columns, 1):
            cell = ws_data.cell(row=1, column=j, value=str(col_name))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        # Rows
        for i, row in df.iterrows():
            for j, val in enumerate(row, 1):
                ws_data.cell(row=i+2, column=j, value=val)
        # Auto width
        for col in ws_data.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws_data.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    # ── Hoja 3: Narrativa IA ──────────────────────────────────
    if narrativa_ia:
        ws_ia = wb.create_sheet("Análisis IA")
        ws_ia["A1"] = "Análisis generado por IA"
        ws_ia["A1"].font = Font(bold=True, size=12)
        ws_ia["A3"] = narrativa_ia
        ws_ia["A3"].alignment = Alignment(wrap_text=True)
        ws_ia.column_dimensions["A"].width = 100
        ws_ia.row_dimensions[3].height = min(len(narrativa_ia) // 3, 400)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
# PDF
# ══════════════════════════════════════════════════════════════

def generar_pdf(df=None, kpis: dict = None, titulo: str = "Dashboard BI",
                periodo: str = "", narrativa_ia: str = "") -> bytes:
    """
    Genera un PDF profesional con:
      - Portada con título y período
      - Tabla de KPIs
      - Narrativa IA (si existe)
      - Top 20 filas del DataFrame (si existe)

    Usa reportlab. Si no está instalado, genera un PDF básico con fpdf2.
    Retorna bytes para st.download_button.
    """
    # Intentar reportlab primero, fallback a fpdf2
    try:
        return _pdf_reportlab(df, kpis, titulo, periodo, narrativa_ia)
    except ImportError:
        try:
            return _pdf_fpdf(df, kpis, titulo, periodo, narrativa_ia)
        except ImportError:
            # Último fallback: PDF mínimo artesanal
            return _pdf_minimo(titulo, periodo, kpis, narrativa_ia)


def _pdf_reportlab(df, kpis, titulo, periodo, narrativa_ia) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    AZUL       = colors.HexColor("#1D3557")
    AZUL_CLARO = colors.HexColor("#EFF6FF")
    GRIS       = colors.HexColor("#64748B")

    # Estilos
    st_titulo = ParagraphStyle("titulo", parent=styles["Title"],
                                fontSize=22, textColor=AZUL, spaceAfter=6,
                                alignment=TA_CENTER)
    st_sub    = ParagraphStyle("sub", parent=styles["Normal"],
                                fontSize=11, textColor=GRIS,
                                alignment=TA_CENTER, spaceAfter=20)
    st_h2     = ParagraphStyle("h2", parent=styles["Heading2"],
                                fontSize=13, textColor=AZUL, spaceBefore=16, spaceAfter=8)
    st_body   = ParagraphStyle("body", parent=styles["Normal"],
                                fontSize=9, leading=14)

    # ── Portada ──
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph(titulo, st_titulo))
    story.append(Paragraph(periodo or datetime.now().strftime("Generado el %d/%m/%Y"), st_sub))
    story.append(HRFlowable(width="100%", thickness=2, color=AZUL))
    story.append(Spacer(1, 0.5*cm))

    # ── KPIs ──
    if kpis:
        story.append(Paragraph("📊 Indicadores Principales", st_h2))
        data = [["Indicador", "Valor"]]
        for k, v in kpis.items():
            if isinstance(v, dict): v = v.get("valor", v)
            data.append([str(k), str(v)])
        tbl = Table(data, colWidths=[10*cm, 6*cm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), AZUL),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, AZUL_CLARO]),
            ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
            ("ALIGN",       (1,0), (1,-1), "RIGHT"),
            ("LEFTPADDING", (0,0), (-1,-1), 8),
            ("RIGHTPADDING",(0,0), (-1,-1), 8),
            ("TOPPADDING",  (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 0.5*cm))

    # ── Narrativa IA ──
    if narrativa_ia:
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#BFDBFE")))
        story.append(Paragraph("🤖 Análisis IA", st_h2))
        # Limpiar markdown básico
        texto_limpio = narrativa_ia.replace("**", "").replace("*", "").replace("#", "")
        for linea in texto_limpio.split("\n"):
            if linea.strip():
                story.append(Paragraph(linea.strip(), st_body))
                story.append(Spacer(1, 3))
        story.append(Spacer(1, 0.5*cm))

    # ── Top datos ──
    if df is not None and not df.empty:
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0")))
        story.append(Paragraph("📋 Top artículos", st_h2))
        df_top   = df.head(20)
        cols_max = min(len(df_top.columns), 6)
        df_show  = df_top.iloc[:, :cols_max]
        data_tbl = [list(df_show.columns)] + df_show.astype(str).values.tolist()
        col_w    = 17*cm / cols_max
        tbl_data = Table(data_tbl, colWidths=[col_w]*cols_max, repeatRows=1)
        tbl_data.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), AZUL),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 7),
            ("ROWBACKGROUNDS", (0,1),(-1,-1),[colors.white, AZUL_CLARO]),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#E2E8F0")),
            ("LEFTPADDING", (0,0), (-1,-1), 4),
            ("RIGHTPADDING",(0,0), (-1,-1), 4),
            ("TOPPADDING",  (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0),(-1,-1), 3),
        ]))
        story.append(tbl_data)

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def _pdf_fpdf(df, kpis, titulo, periodo, narrativa_ia) -> bytes:
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(29, 53, 87)
    pdf.cell(0, 10, titulo, ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 6, periodo or datetime.now().strftime("Generado el %d/%m/%Y"), ln=True, align="C")
    pdf.ln(6)

    if kpis:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(29, 53, 87)
        pdf.cell(0, 8, "Indicadores Principales", ln=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(30, 30, 30)
        for k, v in kpis.items():
            if isinstance(v, dict): v = v.get("valor", v)
            pdf.cell(90, 6, str(k), border=1)
            pdf.cell(90, 6, str(v), border=1, ln=True)
        pdf.ln(4)

    if narrativa_ia:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(29, 53, 87)
        pdf.cell(0, 8, "Análisis IA", ln=True)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(50, 50, 50)
        texto = narrativa_ia.replace("**","").replace("*","")
        pdf.multi_cell(0, 5, texto[:2000])

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf.getvalue()


def _pdf_minimo(titulo, periodo, kpis, narrativa_ia) -> bytes:
    """PDF mínimo sin dependencias externas (solo Python stdlib)."""
    lines = [titulo, periodo or "", ""]
    if kpis:
        lines.append("=== KPIs ===")
        for k, v in (kpis or {}).items():
            if isinstance(v, dict): v = v.get("valor", v)
            lines.append(f"{k}: {v}")
    if narrativa_ia:
        lines.append("")
        lines.append("=== Análisis IA ===")
        lines.append(narrativa_ia[:1000])

    content = "\n".join(lines).encode("latin-1", errors="replace")

    # PDF válido mínimo
    objects = []
    objects.append(b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n")
    objects.append(b"2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n")

    stream = b"BT /F1 10 Tf 50 750 Td 14 TL\n"
    for line in lines[:40]:
        safe = line[:80].encode("latin-1", errors="replace")
        stream += b"(" + safe.replace(b"(", b"\\(").replace(b")", b"\\)") + b") Tj T*\n"
    stream += b"ET"

    objects.append(
        f"3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
        f"/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>\nendobj\n".encode()
    )
    objects.append(
        f"4 0 obj\n<< /Length {len(stream)} >>\nstream\n".encode() +
        stream + b"\nendstream\nendobj\n"
    )
    objects.append(
        b"5 0 obj\n<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>\nendobj\n"
    )

    buf  = io.BytesIO()
    buf.write(b"%PDF-1.4\n")
    offsets = []
    for obj in objects:
        offsets.append(buf.tell())
        buf.write(obj)

    xref_pos = buf.tell()
    buf.write(f"xref\n0 {len(objects)+1}\n".encode())
    buf.write(b"0000000000 65535 f \n")
    for off in offsets:
        buf.write(f"{off:010d} 00000 n \n".encode())
    buf.write(
        f"trailer\n<< /Size {len(objects)+1} /Root 1 0 R >>\n"
        f"startxref\n{xref_pos}\n%%EOF\n".encode()
    )
    buf.seek(0)
    return buf.getvalue()
