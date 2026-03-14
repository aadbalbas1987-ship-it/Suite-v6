"""
robots/robot_cxc.py — RPA Suite v5.7
=======================================
Módulo de Cuentas por Cobrar (CxC).
Procesa archivo Excel de deudores y genera:
  - Reporte aging (0-30, 31-60, 61-90, +90 días)
  - Alertas de mora
  - Resumen por cliente
  - Export Excel con semáforo visual
"""
import os
from pathlib import Path
from datetime import date, datetime

try:
    import pandas as pd
    import numpy as np
    _PD = True
except ImportError:
    _PD = False


# ══════════════════════════════════════════════════════════════
# ESTRUCTURA ESPERADA DEL EXCEL
# Columnas mínimas:
#   CLIENTE, CUIT (opcional), FECHA_VENCIMIENTO, IMPORTE, COMPROBANTE
# ══════════════════════════════════════════════════════════════

COLS_POSIBLES = {
    "cliente":      ["CLIENTE", "Cliente", "RAZON_SOCIAL", "Razon Social", "NOMBRE"],
    "cuit":         ["CUIT", "Cuit", "cuit"],
    "vencimiento":  ["FECHA_VENCIMIENTO", "Vencimiento", "VENC", "Fecha Venc"],
    "importe":      ["IMPORTE", "Importe", "MONTO", "Monto", "SALDO", "Saldo"],
    "comprobante":  ["COMPROBANTE", "Comprobante", "NRO", "Número", "FACTURA"],
    "sucursal":     ["SUCURSAL", "Sucursal", "SUC"],
}


def _col(df: "pd.DataFrame", candidates: list) -> str | None:
    for c in candidates:
        if c in df.columns:
            return c
    return None


def procesar_cxc(
    ruta_excel: str,
    fecha_corte: date = None,
    log_func=None,
) -> dict:
    """
    Procesa el Excel de CxC y retorna un dict con:
      - df_aging:    DataFrame con todas las facturas + días de mora + banda aging
      - resumen:     dict con totales por banda
      - df_clientes: DataFrame agrupado por cliente
      - alertas:     lista de alertas críticas
    """
    def _log(m):
        if log_func: log_func(m)

    if not _PD:
        return {"error": "pandas no disponible"}

    if fecha_corte is None:
        fecha_corte = date.today()

    _log(f"📂 Procesando CxC: {Path(ruta_excel).name}")

    try:
        df = pd.read_excel(ruta_excel, engine="openpyxl")
    except Exception as e:
        return {"error": f"No se pudo leer el archivo: {e}"}

    # Detectar columnas
    col_cli  = _col(df, COLS_POSIBLES["cliente"])
    col_cuit = _col(df, COLS_POSIBLES["cuit"])
    col_venc = _col(df, COLS_POSIBLES["vencimiento"])
    col_imp  = _col(df, COLS_POSIBLES["importe"])
    col_comp = _col(df, COLS_POSIBLES["comprobante"])
    col_suc  = _col(df, COLS_POSIBLES["sucursal"])

    if not col_cli or not col_venc or not col_imp:
        return {"error": f"Columnas requeridas no encontradas. Necesitás: CLIENTE, FECHA_VENCIMIENTO, IMPORTE.\n"
                         f"Columnas detectadas: {list(df.columns)}"}

    # Limpiar y tipificar
    df = df[[c for c in [col_cli, col_cuit, col_venc, col_imp, col_comp, col_suc] if c]].copy()
    df.columns = [c.replace(col_cli, "CLIENTE").replace(col_venc, "VENCIMIENTO")
                  .replace(col_imp, "IMPORTE") for c in df.columns]
    df = df.rename(columns={col_cli: "CLIENTE", col_venc: "VENCIMIENTO", col_imp: "IMPORTE"})
    if col_cuit: df = df.rename(columns={col_cuit: "CUIT"})
    if col_comp: df = df.rename(columns={col_comp: "COMPROBANTE"})
    if col_suc:  df = df.rename(columns={col_suc: "SUCURSAL"})

    df["IMPORTE"] = pd.to_numeric(df["IMPORTE"], errors="coerce").fillna(0)
    df = df[df["IMPORTE"] > 0]  # solo saldos positivos

    df["VENCIMIENTO"] = pd.to_datetime(df["VENCIMIENTO"], errors="coerce", dayfirst=True)
    df = df.dropna(subset=["VENCIMIENTO"])

    # Días de mora
    df["DIAS_MORA"] = (pd.Timestamp(fecha_corte) - df["VENCIMIENTO"]).dt.days.clip(lower=0)

    # Banda aging
    def _banda(dias):
        if dias <= 0:  return "🟢 Al día"
        if dias <= 30: return "🟡 1-30 días"
        if dias <= 60: return "🟠 31-60 días"
        if dias <= 90: return "🔴 61-90 días"
        return "⛔ +90 días"

    df["BANDA"] = df["DIAS_MORA"].apply(_banda)

    # Resumen por banda
    orden_banda = {"🟢 Al día": 0, "🟡 1-30 días": 1, "🟠 31-60 días": 2,
                   "🔴 61-90 días": 3, "⛔ +90 días": 4}
    resumen = {}
    for banda, grp in df.groupby("BANDA"):
        resumen[banda] = {
            "total": round(grp["IMPORTE"].sum(), 2),
            "facturas": len(grp),
            "clientes": grp["CLIENTE"].nunique(),
        }

    total_cartera = df["IMPORTE"].sum()
    total_vencido = df[df["DIAS_MORA"] > 0]["IMPORTE"].sum()

    # Resumen por cliente
    agg = {"IMPORTE": "sum", "DIAS_MORA": "max", "BANDA": lambda x: x.mode()[0] if len(x) else ""}
    if "COMPROBANTE" in df.columns:
        agg["COMPROBANTE"] = "count"
    df_cli = df.groupby("CLIENTE").agg(agg).reset_index()
    df_cli.columns = ["Cliente"] + [
        "Saldo Total", "Días Mora Máx", "Banda Peor"] + (
        ["Facturas"] if "COMPROBANTE" in df.columns else [])
    df_cli = df_cli.sort_values("Saldo Total", ascending=False)

    # Alertas críticas
    alertas = []
    criticos = df[df["DIAS_MORA"] > 90].groupby("CLIENTE")["IMPORTE"].sum()
    for cli, monto in criticos.nlargest(5).items():
        alertas.append(f"⛔ {cli}: ${monto:,.0f} con más de 90 días de mora")
    if total_vencido / total_cartera > 0.3 if total_cartera > 0 else False:
        alertas.append(f"⚠ El {total_vencido/total_cartera*100:.0f}% de la cartera está vencida")

    _log(f"  ✅ CxC procesada: {len(df)} facturas, ${total_cartera:,.0f} total")
    _log(f"  📊 Vencido: ${total_vencido:,.0f} ({total_vencido/total_cartera*100:.1f}%)" if total_cartera else "")
    if alertas:
        for a in alertas:
            _log(f"  {a}")

    return {
        "df_aging":    df.sort_values(["DIAS_MORA", "IMPORTE"], ascending=[False, False]),
        "df_clientes": df_cli,
        "resumen":     resumen,
        "alertas":     alertas,
        "total_cartera":  round(total_cartera, 2),
        "total_vencido":  round(total_vencido, 2),
        "pct_vencido":    round(total_vencido / total_cartera * 100, 1) if total_cartera else 0,
        "fecha_corte":    fecha_corte.isoformat(),
        "n_facturas":     len(df),
        "n_clientes":     df["CLIENTE"].nunique(),
    }


def exportar_cxc_excel(resultado: dict, log_func=None) -> bytes:
    """
    Genera Excel de CxC con 3 hojas:
      - Aging completo
      - Resumen por cliente
      - Resumen ejecutivo
    """
    def _log(m):
        if log_func: log_func(m)

    if not _PD or "error" in resultado:
        return b""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        import io

        wb = Workbook()
        AZUL = PatternFill("solid", fgColor="1D3557")
        VERDE= PatternFill("solid", fgColor="D1FAE5")
        AMARILLO=PatternFill("solid", fgColor="FEF3C7")
        NARANJA =PatternFill("solid", fgColor="FED7AA")
        ROJO =   PatternFill("solid", fgColor="FEE2E2")
        ROJO_F = PatternFill("solid", fgColor="991B1B")
        WHITE_F = Font(color="FFFFFF", bold=True)

        BANDA_FILL = {
            "🟢 Al día":      VERDE,
            "🟡 1-30 días":   AMARILLO,
            "🟠 31-60 días":  NARANJA,
            "🔴 61-90 días":  ROJO,
            "⛔ +90 días":    ROJO_F,
        }

        # ── Hoja 1: Aging ──
        ws1 = wb.active; ws1.title = "Aging Completo"
        df_ag = resultado["df_aging"]
        cols  = list(df_ag.columns)
        for j, c in enumerate(cols, 1):
            cell = ws1.cell(row=1, column=j, value=c)
            cell.fill = AZUL; cell.font = WHITE_F
            cell.alignment = Alignment(horizontal="center")
        for i, (_, row) in enumerate(df_ag.iterrows(), 2):
            fill = BANDA_FILL.get(row.get("BANDA", ""), None)
            for j, val in enumerate(row, 1):
                c = ws1.cell(row=i, column=j, value=val if not hasattr(val, 'isoformat') else str(val)[:10])
                if fill: c.fill = fill
        for col in ws1.columns:
            ws1.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(len(str(c.value or "")) for c in col) + 2, 35)

        # ── Hoja 2: Por cliente ──
        ws2 = wb.create_sheet("Por Cliente")
        df_cl = resultado["df_clientes"]
        for j, c in enumerate(df_cl.columns, 1):
            cell = ws2.cell(row=1, column=j, value=c)
            cell.fill = AZUL; cell.font = WHITE_F
        for i, (_, row) in enumerate(df_cl.iterrows(), 2):
            for j, val in enumerate(row, 1):
                ws2.cell(row=i, column=j, value=val)

        # ── Hoja 3: Resumen ──
        ws3 = wb.create_sheet("Resumen Ejecutivo")
        ws3["A1"] = "Resumen CxC"
        ws3["A1"].font = Font(bold=True, size=14, color="1D3557")
        ws3["A3"] = f"Fecha de corte: {resultado['fecha_corte']}"
        ws3["A4"] = f"Total cartera: ${resultado['total_cartera']:,.2f}"
        ws3["A5"] = f"Total vencido: ${resultado['total_vencido']:,.2f} ({resultado['pct_vencido']}%)"
        ws3["A6"] = f"Facturas: {resultado['n_facturas']} | Clientes: {resultado['n_clientes']}"
        ws3.append([])
        ws3.append(["Banda", "Total $", "Facturas", "Clientes"])
        for col in range(1, 5):
            ws3.cell(row=ws3.max_row, column=col).fill = AZUL
            ws3.cell(row=ws3.max_row, column=col).font = WHITE_F
        for banda, datos in sorted(resultado["resumen"].items(),
                                    key=lambda x: ["🟢","🟡","🟠","🔴","⛔"].index(x[0][0]) if x[0][0] in "🟢🟡🟠🔴⛔" else 9):
            ws3.append([banda, datos["total"], datos["facturas"], datos["clientes"]])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        _log("  ✅ Excel CxC generado")
        return buf.getvalue()

    except Exception as e:
        _log(f"  ❌ Error generando Excel: {e}")
        return b""
